"""
Módulo de dados — todas as operações via Supabase (PostgreSQL na nuvem).
Cada usuário vê apenas os próprios dados via Row Level Security (RLS).
"""

from typing import Optional, List, Dict
from datetime import date
from config import get_client

# Cache em memória: evita re-buscar transações do mesmo mês a cada ação
_tx_cache: Dict[int, List[dict]] = {}


def _invalidate(month_id: int) -> None:
    _tx_cache.pop(month_id, None)
    _inv_net_cache.pop(month_id, None)
    _bill_cache.pop(month_id, None)


def is_cached(month_id: int) -> bool:
    return month_id in _tx_cache


def init_db() -> None:
    get_client()


# ---------------------------------------------------------------------------
# Operações de Meses
# ---------------------------------------------------------------------------

def get_months() -> List[dict]:
    """Retorna os meses do usuário logado, do mais recente ao mais antigo."""
    resp = get_client().table("months") \
        .select("*") \
        .order("year", desc=True) \
        .order("month", desc=True) \
        .execute()
    return resp.data or []


def create_month(name: str, year: int, month: int) -> Optional[dict]:
    """Cria um novo mês para o usuário logado (ignora se já existir)."""
    client  = get_client()
    user_id = client.auth.get_user().user.id

    existing = client.table("months").select("*").eq("name", name).execute()
    if existing.data:
        return existing.data[0]

    resp = client.table("months").insert({
        "name": name, "year": year, "month": month, "user_id": user_id,
    }).execute()
    return resp.data[0] if resp.data else None


def get_month_by_name(name: str) -> Optional[dict]:
    resp = get_client().table("months").select("*").eq("name", name).execute()
    return resp.data[0] if resp.data else None


def rename_month(month_id: int, new_name: str, new_year: int, new_month: int) -> None:
    client  = get_client()
    user_id = client.auth.get_user().user.id
    existing = client.table("months").select("id").eq("name", new_name).eq("user_id", user_id).execute()
    if existing.data and existing.data[0]["id"] != month_id:
        raise ValueError(f'"{new_name}" já existe.')
    client.table("months").update({
        "name": new_name, "year": new_year, "month": new_month,
    }).eq("id", month_id).execute()


def delete_month(month_id: int) -> None:
    get_client().table("months").delete().eq("id", month_id).execute()


# ---------------------------------------------------------------------------
# Operações de Transações
# ---------------------------------------------------------------------------

def get_transactions(month_id: int, tx_type: Optional[str] = None) -> List[dict]:
    if month_id not in _tx_cache:
        resp = get_client().table("transactions") \
            .select("*") \
            .eq("month_id", month_id) \
            .order("id", desc=True) \
            .execute()
        _tx_cache[month_id] = resp.data or []
    txs = _tx_cache[month_id]
    if tx_type:
        return [t for t in txs if t["type"] == tx_type]
    return list(txs)


def add_transaction(
    month_id: int,
    tx_type: str,
    description: str,
    amount: float,
    category: str = "Outros",
    card_id: Optional[int] = None,
    is_expectation: bool = False,
    benefit_id: Optional[int] = None,
    debit_card_id: Optional[int] = None,
    payment_method: Optional[str] = None,
    payment_date: Optional[date] = None,
) -> None:
    client  = get_client()
    user_id = client.auth.get_user().user.id
    row = {
        "month_id":       month_id,
        "user_id":        user_id,
        "type":           tx_type,
        "description":    description,
        "amount":         amount,
        "category":       category,
        "is_expectation": is_expectation,
    }
    if card_id is not None:
        row["card_id"] = card_id
    if benefit_id is not None:
        row["benefit_id"] = benefit_id
    if debit_card_id is not None:
        row["debit_card_id"] = debit_card_id
    if payment_method is not None:
        row["payment_method"] = payment_method
    if payment_date is not None:
        row["payment_date"] = payment_date.isoformat()
    client.table("transactions").insert(row).execute()
    # Gasto pago com benefício debita o saldo na hora (previsões não debitam)
    if benefit_id is not None and not is_expectation:
        _adjust_benefit_balance(benefit_id, -float(amount))
    _invalidate(month_id)


def _find_tx(month_id: int, transaction_id: int) -> Optional[dict]:
    return next((t for t in get_transactions(month_id)
                 if t["id"] == transaction_id), None)


def update_transaction(
    transaction_id: int,
    month_id: int,
    description: str,
    amount: float,
    category: str,
    card_id: Optional[int] = None,
    is_expectation: Optional[bool] = None,
    benefit_id: Optional[int] = None,
    debit_card_id: Optional[int] = None,
    payment_method: Optional[str] = None,
    payment_date: Optional[date] = None,
) -> None:
    # Estorna o efeito antigo no saldo do benefício antes de aplicar o novo
    old = _find_tx(month_id, transaction_id)
    if old and old.get("benefit_id") and not old.get("is_expectation"):
        _adjust_benefit_balance(old["benefit_id"], float(old["amount"] or 0))

    update = {
        "description":    description,
        "amount":         amount,
        "category":       category,
        "card_id":        card_id,
        "benefit_id":     benefit_id,
        "debit_card_id":  debit_card_id,
        "payment_method": payment_method,
        "payment_date":   payment_date.isoformat() if payment_date else None,
    }
    if is_expectation is not None:
        update["is_expectation"] = is_expectation
    get_client().table("transactions").update(update).eq("id", transaction_id).execute()

    debit_now = benefit_id is not None and not (
        is_expectation if is_expectation is not None
        else bool(old and old.get("is_expectation")))
    if debit_now:
        _adjust_benefit_balance(benefit_id, -float(amount))
    _invalidate(month_id)


def import_transactions_bulk(rows: List[dict]) -> None:
    """Confirma uma importação de extrato/fatura: chama add_transaction pra
    cada linha já revisada pelo usuário. Cada dict precisa ter month_id,
    type, description, amount e pode ter category/payment_method/
    payment_date/card_id/benefit_id/debit_card_id. is_expectation é sempre
    False (lançamento importado já é real, nunca previsto)."""
    for r in rows:
        add_transaction(
            month_id=r["month_id"],
            tx_type=r["type"],
            description=r["description"],
            amount=r["amount"],
            category=r.get("category") or "Outros",
            card_id=r.get("card_id"),
            is_expectation=False,
            benefit_id=r.get("benefit_id"),
            debit_card_id=r.get("debit_card_id"),
            payment_method=r.get("payment_method"),
            payment_date=r.get("payment_date"),
        )


def delete_transaction(transaction_id: int, month_id: int) -> None:
    old = _find_tx(month_id, transaction_id)
    if old and old.get("benefit_id") and not old.get("is_expectation"):
        _adjust_benefit_balance(old["benefit_id"], float(old["amount"] or 0))
    get_client().table("transactions").delete().eq("id", transaction_id).execute()
    _invalidate(month_id)


def confirm_expectation(transaction_id: int, month_id: int,
                        description: str, amount: float) -> None:
    """Confirma uma transação prevista: atualiza valor/descrição e marca como real."""
    old = _find_tx(month_id, transaction_id)
    get_client().table("transactions").update({
        "is_expectation": False,
        "description":    description,
        "amount":         amount,
    }).eq("id", transaction_id).execute()
    # Previsão com benefício só debita o saldo ao ser confirmada
    if old and old.get("benefit_id"):
        _adjust_benefit_balance(old["benefit_id"], -float(amount))
    _invalidate(month_id)


def get_total_investments() -> float:
    """Soma líquida de todas as movimentações de investimento do usuário (aportes − saques)."""
    client  = get_client()
    user_id = client.auth.get_user().user.id
    resp = client.table("investment_movements") \
        .select("movement_type,amount") \
        .eq("user_id", user_id) \
        .execute()
    total = 0.0
    for r in (resp.data or []):
        amt = float(r["amount"] or 0)
        if r["movement_type"] == "saque":
            total -= amt
        else:
            total += amt
    return total


def get_month_summary(month_id: int) -> Dict[str, float]:
    rows = get_transactions(month_id)

    real: Dict[str, float] = {
        "entrada_fixa": 0.0, "entrada_variavel": 0.0,
        "saida_fixa":   0.0, "saida_variavel":   0.0,
    }
    proj_extra: Dict[str, float] = {
        "entrada_fixa": 0.0, "entrada_variavel": 0.0,
        "saida_fixa":   0.0, "saida_variavel":   0.0,
    }
    n_expectations = 0
    for row in rows:
        t   = row["type"]
        amt = float(row["amount"] or 0)
        if t not in real:
            continue
        # Compras no cartão não afetam o saldo — só debitam quando a fatura é paga
        if row.get("card_id") and t == "saida_variavel":
            continue
        # Gastos com VR/VA saem do saldo carimbado, não do caixa do mês
        if row.get("benefit_id") and t in ("saida_fixa", "saida_variavel"):
            continue
        if row.get("is_expectation"):
            proj_extra[t] += amt
            n_expectations += 1
        else:
            real[t] += amt

    # Pagamentos de fatura são saídas reais do mês em que foram registrados
    bill_total = sum(float(p["amount"]) for p in get_card_payments(month_id))

    total_entradas      = real["entrada_fixa"] + real["entrada_variavel"]
    total_saidas        = real["saida_fixa"]   + real["saida_variavel"] + bill_total
    total_investimentos = get_month_investment_net(month_id)
    # Investimentos não descontam o saldo — ficam só como informativo
    # (aportar pela aba Investimentos não é a mesma coisa que gastar).
    saldo = total_entradas - total_saidas

    proj_entradas   = total_entradas + proj_extra["entrada_fixa"] + proj_extra["entrada_variavel"]
    proj_saidas     = total_saidas   + proj_extra["saida_fixa"]   + proj_extra["saida_variavel"]
    saldo_projetado = proj_entradas - proj_saidas

    return {
        **real,
        "total_entradas":      total_entradas,
        "total_saidas":        total_saidas,
        "total_investimentos": total_investimentos,
        "saldo":               saldo,
        "saldo_projetado":     saldo_projetado,
        "saldo_acumulado":     get_saldo_acumulado(month_id),
        "n_expectations":      float(n_expectations),
        "has_expectations":    n_expectations > 0,
    }


def _month_real_flow(month_id: int) -> tuple:
    """(entradas, saídas) reais de um mês — mesma regra de get_month_summary
    (ignora previstos, compras no cartão ainda não pagas, gastos em VR/VA),
    usado pelo cálculo do saldo acumulado."""
    rows = get_transactions(month_id)
    entradas = saidas = 0.0
    for row in rows:
        t = row["type"]
        if t not in ("entrada_fixa", "entrada_variavel", "saida_fixa", "saida_variavel"):
            continue
        if row.get("card_id") and t == "saida_variavel":
            continue
        if row.get("benefit_id") and t in ("saida_fixa", "saida_variavel"):
            continue
        if row.get("is_expectation"):
            continue
        amt = float(row["amount"] or 0)
        if t.startswith("entrada"):
            entradas += amt
        else:
            saidas += amt
    bill_total = sum(float(p["amount"]) for p in get_card_payments(month_id))
    saidas += bill_total
    return entradas, saidas


def get_saldo_acumulado(month_id: int) -> float:
    """Saldo real acumulado até (e incluindo) este mês: parte da âncora
    (opening_balance) mais recente em ou antes deste mês — ou do mês mais
    antigo do usuário, com R$0, se nenhuma âncora existir — e soma o fluxo
    real de cada mês em ordem cronológica até o mês alvo."""
    months = get_months()
    target = next((m for m in months if m["id"] == month_id), None)
    if target is None:
        return 0.0

    chronological = sorted(months, key=lambda m: (m["year"], m["month"]))
    target_key = (target["year"], target["month"])

    anchor = None
    for m in chronological:
        if m.get("opening_balance") is None:
            continue
        if (m["year"], m["month"]) <= target_key:
            anchor = m  # o loop está em ordem crescente, então o último achado é o mais recente <= alvo

    if anchor is not None:
        start_balance = float(anchor["opening_balance"])
        start_key = (anchor["year"], anchor["month"])
    elif chronological:
        start_balance = 0.0
        start_key = (chronological[0]["year"], chronological[0]["month"])
    else:
        return 0.0

    total = start_balance
    for m in chronological:
        key = (m["year"], m["month"])
        if start_key <= key <= target_key:
            entradas, saidas = _month_real_flow(m["id"])
            total += entradas - saidas
    return total


def set_month_opening_balance(month_id: int, value) -> None:
    """value=None limpa a âncora desse mês (volta a herdar do mês anterior)."""
    get_client().table("months").update({"opening_balance": value}).eq("id", month_id).execute()


def get_daily_spending(days: int = 7) -> list:
    """Gasto real por dia nos últimos `days` dias (hoje incluso), mesma regra
    de exclusão do resto do app (ignora previstos, compras no cartão ainda
    não pagas, gastos em VR/VA). Não depende do mês/período — olha a data
    real do lançamento (payment_date, ou created_at se não tiver).
    Retorna [{"date": "YYYY-MM-DD", "total": float}, ...] do mais antigo pro
    mais recente, com todos os dias presentes mesmo sem gasto (total 0)."""
    from datetime import date as _date, timedelta

    today  = _date.today()
    window = [today - timedelta(days=i) for i in range(days - 1, -1, -1)]
    totals = {d: 0.0 for d in window}

    resp = get_client().table("transactions").select(
        "type,amount,card_id,benefit_id,is_expectation,payment_date,created_at"
    ).execute()
    for row in (resp.data or []):
        t = row["type"]
        if t not in ("saida_fixa", "saida_variavel"):
            continue
        if row.get("is_expectation"):
            continue
        if row.get("card_id") and t == "saida_variavel":
            continue
        if row.get("benefit_id"):
            continue
        raw = row.get("payment_date") or row.get("created_at")
        try:
            d = _date.fromisoformat(str(raw)[:10])
        except (ValueError, TypeError):
            continue
        if d in totals:
            totals[d] += float(row["amount"] or 0)

    return [{"date": d.isoformat(), "total": totals[d]} for d in window]


def copy_transactions_to_month(from_month_id: int, to_month_id: int) -> int:
    """Copia gastos pós-fechamento de cartão do mês anterior para o novo mês."""
    from datetime import date as _date
    client  = get_client()
    user_id = client.auth.get_user().user.id

    # Mapeamento card_id → closing_day para filtrar pelo ciclo correto
    cards        = get_cards()
    card_closing = {c["id"]: c.get("closing_day", 1) for c in cards}

    rows    = get_transactions(from_month_id)
    to_copy = []
    for r in rows:
        if r["type"] == "saida_variavel" and r.get("card_id") and not r.get("is_expectation"):
            # Só copia se foi adicionado APÓS o fechamento (pertence ao próximo ciclo)
            card_id  = r["card_id"]
            closing  = card_closing.get(card_id, 1)
            raw_date = str(r.get("created_at") or "")[:10]
            try:
                tx_day = _date.fromisoformat(raw_date).day
                if tx_day > closing:
                    to_copy.append(r)
            except ValueError:
                pass  # data inválida → não copia

    for r in to_copy:
        row = {
            "month_id":    to_month_id,
            "user_id":     user_id,
            "type":        r["type"],
            "description": r["description"],
            "amount":      float(r["amount"]),
            "category":    r["category"] or "Outros",
        }
        if r.get("card_id"):
            row["card_id"] = r["card_id"]
        client.table("transactions").insert(row).execute()

    _invalidate(to_month_id)
    return len(to_copy)


def get_expenses_by_category(month_id: int) -> List[dict]:
    rows   = get_transactions(month_id)
    totals: Dict[str, float] = {}
    for row in rows:
        if (row["type"] in ("saida_fixa", "saida_variavel")
                and not row.get("is_expectation")
                and not row.get("benefit_id")):   # carimbado: fora dos envelopes
            cat = row["category"] or "Outros"
            totals[cat] = totals.get(cat, 0.0) + float(row["amount"] or 0)
    result = [{"category": c, "total": t} for c, t in totals.items()]
    result.sort(key=lambda x: x["total"], reverse=True)
    return result


def get_expenses_by_payment_method(month_id: int) -> List[dict]:
    """Igual a get_expenses_by_category, mas agrupado por forma de pagamento.
    Inclui gastos com VR/VA (aqui é uma fatia própria, ao contrário do
    card de categorias, que os exclui por já terem envelope separado)."""
    rows   = get_transactions(month_id)
    totals: Dict[str, float] = {}
    for row in rows:
        if row["type"] in ("saida_fixa", "saida_variavel") and not row.get("is_expectation"):
            pm = row.get("payment_method") or "outro"
            totals[pm] = totals.get(pm, 0.0) + float(row["amount"] or 0)
    result = [{"payment_method": pm, "total": t} for pm, t in totals.items()]
    result.sort(key=lambda x: x["total"], reverse=True)
    return result


# ---------------------------------------------------------------------------
# Operações de Metas
# ---------------------------------------------------------------------------

def get_goals() -> List[dict]:
    resp = get_client().table("goals").select("*").order("created_at", desc=False).execute()
    return resp.data or []


def create_goal(name: str, target_amount: float) -> Optional[dict]:
    client  = get_client()
    user_id = client.auth.get_user().user.id
    resp = client.table("goals").insert({
        "name": name, "target_amount": target_amount,
        "saved_amount": 0, "user_id": user_id,
    }).execute()
    return resp.data[0] if resp.data else None


def add_goal_contribution(goal_id: int, amount: float) -> None:
    resp = get_client().table("goals").select("saved_amount").eq("id", goal_id).execute()
    if resp.data:
        current = float(resp.data[0]["saved_amount"] or 0)
        get_client().table("goals").update(
            {"saved_amount": max(0.0, current + amount)}
        ).eq("id", goal_id).execute()


def update_goal(goal_id: int, name: str, target_amount: float) -> None:
    get_client().table("goals").update({
        "name": name, "target_amount": target_amount,
    }).eq("id", goal_id).execute()


def delete_goal(goal_id: int) -> None:
    get_client().table("goals").delete().eq("id", goal_id).execute()


# ---------------------------------------------------------------------------
# Cartões de Crédito
# ---------------------------------------------------------------------------

_card_tx_cache: Dict[str, list] = {}  # key: "{card_id}_{month_id}"
_inv_net_cache: Dict[int, float] = {}  # key: month_id → net investment do mês
_bill_cache:    Dict[int, list]  = {}  # key: month_id → credit_card_payments


def get_cards() -> List[dict]:
    resp = get_client().table("credit_cards").select("*").order("created_at").execute()
    return resp.data or []


def create_card(name: str, limit: float, due_day: int, closing_day: int, color: str) -> Optional[dict]:
    client  = get_client()
    user_id = client.auth.get_user().user.id
    resp = client.table("credit_cards").insert({
        "name": name, "limit": limit, "due_day": due_day,
        "closing_day": closing_day, "color": color, "user_id": user_id,
    }).execute()
    return resp.data[0] if resp.data else None


def update_card(card_id: int, name: str, limit: float, due_day: int, closing_day: int, color: str) -> None:
    get_client().table("credit_cards").update({
        "name": name, "limit": limit, "due_day": due_day,
        "closing_day": closing_day, "color": color,
    }).eq("id", card_id).execute()


def delete_card(card_id: int) -> None:
    get_client().table("credit_cards").delete().eq("id", card_id).execute()
    # Limpa cache de transações deste cartão
    keys = [k for k in _card_tx_cache if k.startswith(f"{card_id}_")]
    for k in keys:
        _card_tx_cache.pop(k, None)


def get_card_transactions(card_id: int, month_id: int) -> List[dict]:
    key = f"{card_id}_{month_id}"
    if key not in _card_tx_cache:
        resp = get_client().table("card_transactions") \
            .select("*") \
            .eq("card_id", card_id) \
            .eq("month_id", month_id) \
            .order("id", desc=True) \
            .execute()
        _card_tx_cache[key] = resp.data or []
    return list(_card_tx_cache[key])


def add_card_transaction(card_id: int, month_id: int, description: str, amount: float) -> None:
    client  = get_client()
    user_id = client.auth.get_user().user.id
    client.table("card_transactions").insert({
        "card_id": card_id, "month_id": month_id,
        "description": description, "amount": amount, "user_id": user_id,
    }).execute()
    _card_tx_cache.pop(f"{card_id}_{month_id}", None)


def delete_card_transaction(tx_id: int, card_id: int, month_id: int) -> None:
    get_client().table("card_transactions").delete().eq("id", tx_id).execute()
    _card_tx_cache.pop(f"{card_id}_{month_id}", None)


def get_card_payments(month_id: int) -> List[dict]:
    """Retorna os pagamentos de fatura registrados no mês."""
    if month_id not in _bill_cache:
        resp = get_client().table("credit_card_payments") \
            .select("*") \
            .eq("month_id", month_id) \
            .execute()
        _bill_cache[month_id] = resp.data or []
    return list(_bill_cache[month_id])


def pay_card_bill(card_id: int, month_id: int, amount: float, note: str = "") -> None:
    """Registra o pagamento de uma fatura de cartão."""
    client  = get_client()
    user_id = client.auth.get_user().user.id
    client.table("credit_card_payments").insert({
        "card_id":  card_id,
        "month_id": month_id,
        "amount":   amount,
        "note":     note or None,
        "user_id":  user_id,
    }).execute()
    _bill_cache.pop(month_id, None)


def _card_cycle_start(closing_day: int):
    """Início do ciclo de faturamento atual (réplica de ui.credit_cards._cycle_start)."""
    import calendar
    from datetime import date
    today = date.today()
    if today.day >= closing_day:
        try:
            return date(today.year, today.month, closing_day)
        except ValueError:
            return date(today.year, today.month, 1)
    if today.month == 1:
        y, m = today.year - 1, 12
    else:
        y, m = today.year, today.month - 1
    max_day = calendar.monthrange(y, m)[1]
    return date(y, m, min(closing_day, max_day))


def settle_card_bill(card_id: int, month_id: int, closing_day: int,
                     card_name: str) -> float:
    """Quita a fatura: soma as compras reais do ciclo, exclui esses lançamentos
    e cria uma única saída 'Pagamento fatura cartão de crédito' que debita o
    saldo. Retorna o valor pago (0.0 se não havia fatura em aberto)."""
    from datetime import date
    client  = get_client()
    user_id = client.auth.get_user().user.id
    start   = _card_cycle_start(closing_day)

    to_settle = []
    for tx in get_transactions(month_id):
        if (tx.get("card_id") != card_id or tx["type"] != "saida_variavel"
                or tx.get("is_expectation")):
            continue
        raw = str(tx.get("created_at") or "")[:10]
        try:
            if date.fromisoformat(raw) >= start:
                to_settle.append(tx)
        except ValueError:
            to_settle.append(tx)

    total = sum(float(t["amount"] or 0) for t in to_settle)
    if total <= 0:
        return 0.0

    for t in to_settle:
        client.table("transactions").delete().eq("id", t["id"]).execute()

    client.table("transactions").insert({
        "month_id":    month_id,
        "user_id":     user_id,
        "type":        "saida_variavel",
        "description": f"Pagamento fatura cartão de crédito — {card_name}",
        "amount":      total,
        "category":    "Outros",
    }).execute()

    _invalidate(month_id)
    return total


# ---------------------------------------------------------------------------
# Cartões de Débito
# ---------------------------------------------------------------------------
# Entidade simples (só nome/cor): débito não tem fatura/limite/vencimento —
# a transação já debita o saldo na hora, igual a "Nenhuma origem".

def get_debit_cards() -> List[dict]:
    """Usado só pelo seletor de forma de pagamento (não há mais UI pra
    cadastrar cartão de débito novo — só lista os que já existem)."""
    resp = get_client().table("debit_cards").select("*").order("created_at").execute()
    return resp.data or []


def clear_cache() -> None:
    """Limpa todo o cache (chamado no logout)."""
    _tx_cache.clear()
    _card_tx_cache.clear()
    _inv_net_cache.clear()
    _bill_cache.clear()
    _plan_cache.clear()
    _plan_items_cache.clear()
    _invalidate_debts()
    _invalidate_benefits()


# ---------------------------------------------------------------------------
# Investimentos
# ---------------------------------------------------------------------------

def get_month_investment_net(month_id: int) -> float:
    """Retorna o valor líquido de investimentos do mês (aportes − saques). Usa cache."""
    if month_id not in _inv_net_cache:
        resp = get_client().table("investment_movements") \
            .select("movement_type,amount") \
            .eq("month_id", month_id) \
            .execute()
        total = 0.0
        for r in (resp.data or []):
            amt = float(r["amount"] or 0)
            if r["movement_type"] == "saque":
                total -= amt
            else:
                total += amt
        _inv_net_cache[month_id] = total
    return _inv_net_cache[month_id]


def get_investments(include_archived: bool = False) -> List[dict]:
    """Retorna todos os investimentos do usuário. Por padrão exclui arquivados."""
    query = get_client().table("investments") \
        .select("*") \
        .order("created_at", desc=False)
    if not include_archived:
        query = query.is_("archived_at", "null")
    return query.execute().data or []


def create_investment(
    name: str,
    category: str,
    month_id: int,
    amount: float,
    note: str = "",
) -> dict:
    """Cria a entidade de investimento e registra o aporte inicial."""
    client  = get_client()
    user_id = client.auth.get_user().user.id

    inv_resp = client.table("investments").insert({
        "user_id":  user_id,
        "name":     name,
        "category": category,
    }).execute()
    if not inv_resp.data:
        raise RuntimeError("Falha ao criar investimento.")
    inv = inv_resp.data[0]

    client.table("investment_movements").insert({
        "investment_id": inv["id"],
        "user_id":       user_id,
        "month_id":      month_id,
        "movement_type": "aporte_inicial",
        "amount":        amount,
        "note":          note or None,
    }).execute()
    _inv_net_cache.pop(month_id, None)
    return inv


def get_investment_movements(investment_id: int) -> List[dict]:
    """Retorna todas as movimentações de um investimento, mais recente primeiro."""
    resp = get_client().table("investment_movements") \
        .select("*") \
        .eq("investment_id", investment_id) \
        .order("created_at", desc=True) \
        .execute()
    return resp.data or []


def add_movement(
    investment_id: int,
    month_id: int,
    movement_type: str,
    amount: float,
    note: str = "",
) -> None:
    """Adiciona um aporte ou saque a um investimento existente."""
    client  = get_client()
    user_id = client.auth.get_user().user.id
    client.table("investment_movements").insert({
        "investment_id": investment_id,
        "user_id":       user_id,
        "month_id":      month_id,
        "movement_type": movement_type,
        "amount":        amount,
        "note":          note or None,
    }).execute()
    _inv_net_cache.pop(month_id, None)


def update_investment(investment_id: int, name: str, category: str) -> None:
    get_client().table("investments").update({
        "name": name, "category": category,
    }).eq("id", investment_id).execute()


def update_movement(movement_id: int, amount: float, note: str) -> None:
    get_client().table("investment_movements").update({
        "amount": amount, "note": note or None,
    }).eq("id", movement_id).execute()
    _inv_net_cache.clear()


def delete_movement(movement_id: int) -> None:
    get_client().table("investment_movements").delete().eq("id", movement_id).execute()
    _inv_net_cache.clear()


def archive_investment(investment_id: int) -> None:
    """Arquiva o investimento (some da lista mas mantém saldo nos totais)."""
    from datetime import datetime, timezone
    get_client().table("investments").update({
        "archived_at": datetime.now(timezone.utc).isoformat(),
    }).eq("id", investment_id).execute()


def delete_investment(investment_id: int) -> None:
    """Exclui permanentemente o investimento e todas as suas movimentações."""
    client = get_client()
    client.table("investment_movements").delete().eq("investment_id", investment_id).execute()
    client.table("investments").delete().eq("id", investment_id).execute()
    _inv_net_cache.clear()


def get_all_investment_movements() -> List[dict]:
    """Retorna todas as movimentações do usuário (para cálculo de saldos em lote)."""
    client  = get_client()
    user_id = client.auth.get_user().user.id
    resp = client.table("investment_movements") \
        .select("*") \
        .eq("user_id", user_id) \
        .order("created_at", desc=True) \
        .execute()
    return resp.data or []


# ---------------------------------------------------------------------------
# Planejamento Mensal
# ---------------------------------------------------------------------------

_plan_cache:       Dict[int, Optional[dict]] = {}  # month_id → plano (ou None)
_plan_items_cache: Dict[int, List[dict]]     = {}  # plan_id  → itens


def get_plan(month_id: int) -> Optional[dict]:
    """Retorna o plano do mês, ou None se ainda não existir."""
    if month_id not in _plan_cache:
        resp = get_client().table("monthly_plans") \
            .select("*") \
            .eq("month_id", month_id) \
            .execute()
        _plan_cache[month_id] = resp.data[0] if resp.data else None
    return _plan_cache[month_id]


def get_plan_items(plan_id: int) -> List[dict]:
    """Retorna os itens (alocações por categoria) de um plano."""
    if plan_id not in _plan_items_cache:
        resp = get_client().table("monthly_plan_items") \
            .select("*") \
            .eq("plan_id", plan_id) \
            .order("planned_amount", desc=True) \
            .execute()
        _plan_items_cache[plan_id] = resp.data or []
    return list(_plan_items_cache[plan_id])


def save_plan(month_id: int, income: float, items: List[dict]) -> dict:
    """Cria ou atualiza o plano do mês e substitui seus itens.

    Nunca duplica: se o plano já existe (constraint unique month_id),
    edita o existente. Ao criar um plano novo, fecha automaticamente
    planos `ativo` de meses anteriores.

    items: [{"category", "planned_amount", "suggested_amount", "is_eventual"}]
    """
    from datetime import datetime, timezone
    client  = get_client()
    user_id = client.auth.get_user().user.id
    now_iso = datetime.now(timezone.utc).isoformat()

    _plan_cache.pop(month_id, None)
    plan   = get_plan(month_id)
    is_new = plan is None

    if is_new:
        try:
            resp = client.table("monthly_plans").insert({
                "month_id": month_id, "user_id": user_id, "income": income,
            }).execute()
            plan = resp.data[0]
        except Exception:
            # Plano criado em paralelo (unique month_id) → edita o existente
            _plan_cache.pop(month_id, None)
            plan = get_plan(month_id)
            if plan is None:
                raise
            is_new = False

    if not is_new:
        client.table("monthly_plans").update({
            "income": income, "updated_at": now_iso,
        }).eq("id", plan["id"]).execute()
        client.table("monthly_plan_items").delete().eq("plan_id", plan["id"]).execute()
    else:
        _close_previous_plans(month_id)

    rows = [{
        "plan_id":          plan["id"],
        "user_id":          user_id,
        "category":         it["category"],
        "planned_amount":   float(it.get("planned_amount") or 0),
        "suggested_amount": it.get("suggested_amount"),
        "is_eventual":      bool(it.get("is_eventual")),
        "is_mandatory":     bool(it.get("is_mandatory")),
    } for it in items]
    if rows:
        client.table("monthly_plan_items").insert(rows).execute()

    _plan_cache.pop(month_id, None)
    _plan_items_cache.pop(plan["id"], None)
    return plan


def _close_previous_plans(month_id: int) -> None:
    """Fecha planos `ativo` de meses anteriores ao mês dado."""
    months  = get_months()
    current = next((m for m in months if m["id"] == month_id), None)
    if current is None:
        return
    key       = (current["year"], current["month"])
    prior_ids = [m["id"] for m in months if (m["year"], m["month"]) < key]
    if not prior_ids:
        return
    get_client().table("monthly_plans") \
        .update({"status": "fechado"}) \
        .eq("status", "ativo") \
        .in_("month_id", prior_ids) \
        .execute()
    for mid in prior_ids:
        _plan_cache.pop(mid, None)


def get_plan_history(month_id: int, n: int = 3) -> tuple:
    """Histórico para a sugestão do plano: até n meses anteriores ao mês dado.

    Retorna (gastos, rendas), ambos do mês mais recente ao mais antigo:
      gastos: [{categoria: total}, ...]
      rendas: [total_entradas, ...]  (apenas entradas confirmadas)
    """
    months  = get_months()  # já vem ordenado do mais recente ao mais antigo
    current = next((m for m in months if m["id"] == month_id), None)
    if current is None:
        return [], []
    key   = (current["year"], current["month"])
    prior = [m for m in months if (m["year"], m["month"]) < key][:n]

    expenses, incomes = [], []
    for m in prior:
        cat_totals = {r["category"]: float(r["total"]) for r in get_expenses_by_category(m["id"])}
        inv_net = get_month_investment_net(m["id"])
        if inv_net > 0:
            cat_totals["Investimentos"] = inv_net
        expenses.append(cat_totals)
        incomes.append(get_month_income(m["id"]))
    return expenses, incomes


def get_plan_realized(month_id: int) -> Dict[str, float]:
    """Realizado por categoria para o plano: gastos + aportes líquidos do mês."""
    realized = {r["category"]: float(r["total"]) for r in get_expenses_by_category(month_id)}
    inv_net = get_month_investment_net(month_id)
    if inv_net > 0:
        realized["Investimentos"] = inv_net
    return realized


def get_month_income(month_id: int, include_expectations: bool = False) -> float:
    """Soma das entradas do mês (fixas + variáveis)."""
    total = 0.0
    for r in get_transactions(month_id):
        if r["type"] not in ("entrada_fixa", "entrada_variavel"):
            continue
        if r.get("is_expectation") and not include_expectations:
            continue
        total += float(r["amount"] or 0)
    return total


# ---------------------------------------------------------------------------
# Dívidas
# ---------------------------------------------------------------------------

_debts_cache: Optional[List[dict]] = None
_inst_cache:  Optional[List[dict]] = None


def _invalidate_debts() -> None:
    global _debts_cache, _inst_cache
    _debts_cache = None
    _inst_cache  = None


def get_debts() -> List[dict]:
    global _debts_cache
    if _debts_cache is None:
        resp = get_client().table("debts").select("*") \
            .order("created_at", desc=False).execute()
        _debts_cache = resp.data or []
    return list(_debts_cache)


def get_all_installments() -> List[dict]:
    global _inst_cache
    if _inst_cache is None:
        resp = get_client().table("debt_installments").select("*") \
            .order("due_year").order("due_month").order("installment_number") \
            .execute()
        _inst_cache = resp.data or []
    return list(_inst_cache)


def installment_status(inst: dict) -> str:
    """Status derivado: paga | atrasada | pendente (nada é gravado no banco)."""
    from datetime import date
    if inst.get("paid_at"):
        return "paga"
    today = date.today()
    if (inst["due_year"], inst["due_month"]) < (today.year, today.month):
        return "atrasada"
    return "pendente"


def create_debt(description: str, creditor: str, total_amount: float,
                category: str, notes: str, installments: List[dict]) -> dict:
    """Cria a dívida e suas parcelas.

    installments: [{"number", "amount", "year", "month"}]
    """
    client  = get_client()
    user_id = client.auth.get_user().user.id
    resp = client.table("debts").insert({
        "user_id":      user_id,
        "description":  description,
        "creditor":     creditor or None,
        "total_amount": total_amount,
        "category":     category or "Dívidas",
        "notes":        notes or None,
    }).execute()
    debt = resp.data[0]
    rows = [{
        "debt_id":            debt["id"],
        "user_id":            user_id,
        "installment_number": it["number"],
        "amount":             float(it["amount"]),
        "due_year":           int(it["year"]),
        "due_month":          int(it["month"]),
    } for it in installments]
    client.table("debt_installments").insert(rows).execute()
    _invalidate_debts()
    return debt


def update_debt(debt_id: int, description: str, creditor: str,
                category: str, notes: str) -> None:
    get_client().table("debts").update({
        "description": description,
        "creditor":    creditor or None,
        "category":    category or "Dívidas",
        "notes":       notes or None,
    }).eq("id", debt_id).execute()
    _invalidate_debts()


def update_installment_amount(inst_id: int, debt_id: int, amount: float) -> None:
    """Atualiza o valor de uma parcela e recalcula o total da dívida."""
    client = get_client()
    client.table("debt_installments").update({"amount": amount}) \
        .eq("id", inst_id).execute()
    _invalidate_debts()
    total = sum(float(i["amount"]) for i in get_all_installments()
                if i["debt_id"] == debt_id)
    client.table("debts").update({"total_amount": total}).eq("id", debt_id).execute()
    _invalidate_debts()


def reschedule_installment(inst_id: int, year: int, month: int) -> None:
    get_client().table("debt_installments").update({
        "due_year": year, "due_month": month,
    }).eq("id", inst_id).execute()
    _invalidate_debts()


def pay_installment(inst: dict, debt: dict, n_total: int,
                    launch_expense: bool) -> None:
    """Marca a parcela como paga; opcionalmente lança o gasto no mês dela."""
    from datetime import datetime, timezone
    client  = get_client()
    user_id = client.auth.get_user().user.id

    expense_id = None
    if launch_expense:
        month = _ensure_month(inst["due_year"], inst["due_month"])
        desc  = debt["description"]
        if n_total > 1:
            desc += f" (parcela {inst['installment_number']}/{n_total})"
        resp = client.table("transactions").insert({
            "month_id":    month["id"],
            "user_id":     user_id,
            "type":        "saida_fixa",
            "description": desc,
            "amount":      float(inst["amount"]),
            "category":    debt.get("category") or "Dívidas",
        }).execute()
        if resp.data:
            expense_id = resp.data[0]["id"]
        _invalidate(month["id"])

    client.table("debt_installments").update({
        "paid_at":    datetime.now(timezone.utc).isoformat(),
        "expense_id": expense_id,
    }).eq("id", inst["id"]).execute()
    _invalidate_debts()


def undo_payment(inst: dict) -> None:
    """Desfaz o pagamento; remove o gasto vinculado, se houver."""
    client = get_client()
    if inst.get("expense_id"):
        client.table("transactions").delete().eq("id", inst["expense_id"]).execute()
        _tx_cache.clear()
    client.table("debt_installments").update({
        "paid_at": None, "expense_id": None,
    }).eq("id", inst["id"]).execute()
    _invalidate_debts()


def delete_installment(inst: dict, delete_expense: bool = False) -> None:
    client = get_client()
    if delete_expense and inst.get("expense_id"):
        client.table("transactions").delete().eq("id", inst["expense_id"]).execute()
        _tx_cache.clear()
    client.table("debt_installments").delete().eq("id", inst["id"]).execute()
    _invalidate_debts()
    # Recalcula o total da dívida (ou remove a dívida se ficou sem parcelas)
    remaining = [i for i in get_all_installments() if i["debt_id"] == inst["debt_id"]]
    if remaining:
        total = sum(float(i["amount"]) for i in remaining)
        client.table("debts").update({"total_amount": total}) \
            .eq("id", inst["debt_id"]).execute()
    else:
        client.table("debts").delete().eq("id", inst["debt_id"]).execute()
    _invalidate_debts()


def delete_debt(debt_id: int, delete_expenses: bool = False) -> None:
    client = get_client()
    if delete_expenses:
        ids = [i["expense_id"] for i in get_all_installments()
               if i["debt_id"] == debt_id and i.get("expense_id")]
        for eid in ids:
            client.table("transactions").delete().eq("id", eid).execute()
        if ids:
            _tx_cache.clear()
    client.table("debts").delete().eq("id", debt_id).execute()
    _invalidate_debts()


def _ensure_month(year: int, month: int) -> dict:
    """Retorna o período (year, month), criando-o se ainda não existir."""
    from utils.helpers import month_name_from_num
    name = month_name_from_num(month, year)
    existing = get_month_by_name(name)
    if existing:
        return existing
    created = create_month(name, year, month)
    if created is None:
        raise RuntimeError(f"Falha ao criar o período {name}.")
    return created


def get_month_debt_totals(year: int, month: int) -> Dict[str, float]:
    """Parcelas não pagas com vencimento no mês, somadas por categoria."""
    debts_by_id = {d["id"]: d for d in get_debts()}
    totals: Dict[str, float] = {}
    for i in get_all_installments():
        if i.get("paid_at") or i["due_year"] != year or i["due_month"] != month:
            continue
        cat = debts_by_id.get(i["debt_id"], {}).get("category") or "Dívidas"
        totals[cat] = totals.get(cat, 0.0) + float(i["amount"])
    return totals


def get_month_debt_totals_for(month_id: int) -> Dict[str, float]:
    months  = get_months()
    current = next((m for m in months if m["id"] == month_id), None)
    if current is None:
        return {}
    return get_month_debt_totals(current["year"], current["month"])


def sync_debts_into_plan(month_id: int) -> bool:
    """Garante que o plano ativo do mês reflita as parcelas pendentes.

    Soma as parcelas não pagas do mês por categoria e faz upsert nos itens
    do plano com is_mandatory=true; remove itens obrigatórios cuja dívida
    sumiu. Retorna True se o plano foi alterado.
    """
    plan = get_plan(month_id)
    if not plan or plan.get("status") != "ativo":
        return False
    desired = get_month_debt_totals_for(month_id)
    items   = get_plan_items(plan["id"])

    client  = get_client()
    user_id = client.auth.get_user().user.id
    by_cat  = {i["category"]: i for i in items}
    changed = False

    for cat, total in desired.items():
        cur = by_cat.get(cat)
        if cur is None:
            client.table("monthly_plan_items").insert({
                "plan_id":          plan["id"],
                "user_id":          user_id,
                "category":         cat,
                "planned_amount":   total,
                "suggested_amount": total,
                "is_mandatory":     True,
            }).execute()
            changed = True
        elif (not cur.get("is_mandatory")
              or abs(float(cur["planned_amount"] or 0) - total) > 0.005):
            client.table("monthly_plan_items").update({
                "planned_amount":   total,
                "suggested_amount": total,
                "is_mandatory":     True,
            }).eq("id", cur["id"]).execute()
            changed = True

    for cat, item in by_cat.items():
        if item.get("is_mandatory") and cat not in desired:
            client.table("monthly_plan_items").delete() \
                .eq("id", item["id"]).execute()
            changed = True

    if changed:
        _plan_items_cache.pop(plan["id"], None)
    return changed


def get_debt_overview() -> dict:
    """Resumo: total em aberto, nº de atrasadas e comprometimento futuro (6 meses)."""
    from datetime import date
    insts = get_all_installments()
    total_aberto = 0.0
    n_atrasadas  = 0
    for i in insts:
        st = installment_status(i)
        if st != "paga":
            total_aberto += float(i["amount"])
        if st == "atrasada":
            n_atrasadas += 1

    today = date.today()
    future = []
    y, m = today.year, today.month
    for _ in range(6):
        total = sum(float(i["amount"]) for i in insts
                    if not i.get("paid_at")
                    and i["due_year"] == y and i["due_month"] == m)
        future.append({"year": y, "month": m, "total": total})
        m += 1
        if m > 12:
            m = 1
            y += 1
    return {
        "total_aberto": total_aberto,
        "n_atrasadas":  n_atrasadas,
        "future":       future,
    }


# ---------------------------------------------------------------------------
# Benefícios (VR / VA)
# ---------------------------------------------------------------------------

_benefits_cache: Optional[List[dict]] = None


def _invalidate_benefits() -> None:
    global _benefits_cache
    _benefits_cache = None


def get_benefits(include_archived: bool = False) -> List[dict]:
    global _benefits_cache
    if _benefits_cache is None:
        resp = get_client().table("benefit_cards").select("*") \
            .order("created_at", desc=False).execute()
        _benefits_cache = resp.data or []
    if include_archived:
        return list(_benefits_cache)
    return [b for b in _benefits_cache if not b.get("archived_at")]


def get_benefit_by_id(benefit_id: int) -> Optional[dict]:
    return next((b for b in get_benefits(include_archived=True)
                 if b["id"] == benefit_id), None)


def _clamp_day(year: int, month: int, day: int) -> int:
    import calendar
    return min(day, calendar.monthrange(year, month)[1])


def _renewal_date(year: int, month: int, renewal_day: int):
    from datetime import date
    return date(year, month, _clamp_day(year, month, renewal_day))


def _last_occurrence(renewal_day: int, ref):
    """Data de renovação mais recente <= ref (dia ajustado p/ meses curtos)."""
    d = _renewal_date(ref.year, ref.month, renewal_day)
    if d <= ref:
        return d
    if ref.month == 1:
        y, m = ref.year - 1, 12
    else:
        y, m = ref.year, ref.month - 1
    return _renewal_date(y, m, renewal_day)


def days_until_renewal(benefit: dict) -> int:
    from datetime import date
    today = date.today()
    d = _renewal_date(today.year, today.month, benefit["renewal_day"])
    if d <= today:
        if today.month == 12:
            y, m = today.year + 1, 1
        else:
            y, m = today.year, today.month + 1
        d = _renewal_date(y, m, benefit["renewal_day"])
    return (d - today).days


def create_benefit(name: str, benefit_type: str, balance: float,
                   renewal_day: int, recharge_amount: float,
                   recharge_mode: str, color: str) -> dict:
    """Cria o benefício. O saldo inicial NÃO gera registro de renovação.

    last_renewal é fixado na última renovação já ocorrida, para que a próxima
    abertura do app não dispare uma renovação retroativa sobre o saldo inicial.
    """
    from datetime import date
    client  = get_client()
    user_id = client.auth.get_user().user.id
    last = _last_occurrence(renewal_day, date.today()).isoformat()
    resp = client.table("benefit_cards").insert({
        "user_id":         user_id,
        "name":            name,
        "benefit_type":    benefit_type,
        "balance":         balance,
        "renewal_day":     renewal_day,
        "recharge_amount": recharge_amount,
        "recharge_mode":   recharge_mode,
        "color":           color,
        "last_renewal":    last,
    }).execute()
    _invalidate_benefits()
    return resp.data[0]


def update_benefit(benefit_id: int, name: str, benefit_type: str,
                   renewal_day: int, recharge_amount: float,
                   recharge_mode: str, color: str) -> None:
    """Edita o benefício. Mudanças de recarga/dia valem a partir da próxima
    renovação (sem efeito retroativo). O saldo não é alterado aqui."""
    get_client().table("benefit_cards").update({
        "name":            name,
        "benefit_type":    benefit_type,
        "renewal_day":     renewal_day,
        "recharge_amount": recharge_amount,
        "recharge_mode":   recharge_mode,
        "color":           color,
    }).eq("id", benefit_id).execute()
    _invalidate_benefits()


def set_benefit_balance(benefit_id: int, balance: float) -> None:
    """Ajuste manual de saldo (correção)."""
    get_client().table("benefit_cards").update(
        {"balance": balance}).eq("id", benefit_id).execute()
    _invalidate_benefits()


def archive_benefit(benefit_id: int) -> None:
    """Exclusão = arquivamento: some da lista mas mantém os gastos vinculados."""
    from datetime import datetime, timezone
    get_client().table("benefit_cards").update({
        "archived_at": datetime.now(timezone.utc).isoformat(),
    }).eq("id", benefit_id).execute()
    _invalidate_benefits()


def get_benefit_renewals(benefit_id: int) -> List[dict]:
    resp = get_client().table("benefit_renewals").select("*") \
        .eq("benefit_id", benefit_id) \
        .order("renewed_at", desc=True).execute()
    return resp.data or []


def _adjust_benefit_balance(benefit_id: int, delta: float) -> None:
    """delta < 0 debita (gasto); delta > 0 credita (estorno)."""
    b = get_benefit_by_id(benefit_id)
    if b is None:
        return
    new_balance = float(b["balance"] or 0) + delta
    get_client().table("benefit_cards").update(
        {"balance": new_balance}).eq("id", benefit_id).execute()
    _invalidate_benefits()


def apply_due_renewals(benefit: dict) -> List[dict]:
    """Aplica todas as renovações pendentes entre last_renewal e hoje.

    Modo 'acumula': saldo += recarga a cada renovação.
    Modo 'zera':    saldo = recarga a cada renovação (na prática só a última
                    muda o saldo, mas todas são registradas no histórico).
    Retorna a lista de renovações aplicadas (para feedback na UI).
    """
    from datetime import date
    raw = benefit.get("last_renewal")
    if not raw:
        return []
    last = date.fromisoformat(str(raw)[:10])
    today = date.today()
    recharge = float(benefit.get("recharge_amount") or 0)
    mode     = benefit.get("recharge_mode") or "acumula"
    balance  = float(benefit.get("balance") or 0)

    client  = get_client()
    user_id = client.auth.get_user().user.id

    applied   = []
    rows       = []
    y, m = last.year, last.month
    while True:
        if m == 12:
            y, m = y + 1, 1
        else:
            m += 1
        nxt = _renewal_date(y, m, benefit["renewal_day"])
        if nxt > today:
            break
        before  = balance
        balance = recharge if mode == "zera" else balance + recharge
        rows.append({
            "benefit_id":     benefit["id"],
            "user_id":        user_id,
            "renewed_at":     nxt.isoformat(),
            "amount":         recharge,
            "balance_before": before,
            "balance_after":  balance,
        })
        applied.append({"date": nxt, "amount": recharge,
                        "balance_after": balance})
        last = nxt

    if rows:
        client.table("benefit_renewals").insert(rows).execute()
        client.table("benefit_cards").update({
            "balance":      balance,
            "last_renewal": last.isoformat(),
        }).eq("id", benefit["id"]).execute()
        _invalidate_benefits()
    return applied


def apply_all_due_renewals() -> List[dict]:
    """Roda na abertura do app. Retorna resumo p/ feedback:
    [{"name", "benefit_type", "total", "count", "balance_after"}]."""
    summary = []
    for b in get_benefits():
        applied = apply_due_renewals(b)
        if applied:
            summary.append({
                "name":          b["name"],
                "benefit_type":  b["benefit_type"],
                "total":         sum(a["amount"] for a in applied),
                "count":         len(applied),
                "balance_after": applied[-1]["balance_after"],
            })
    return summary


_EXPORT_TYPE_LABELS = {
    "entrada_fixa":     "Entrada Fixa",
    "entrada_variavel": "Entrada Variável",
    "saida_fixa":       "Saída Fixa",
    "saida_variavel":   "Saída Variável",
    "investimento":     "Investimento",
}


def export_month_xlsx(month_id: int, month_name: str):
    """Monta um Workbook (openpyxl) formatado com os lançamentos do mês —
    cabeçalho destacado, valores como número (não texto), entradas em verde
    e saídas em vermelho, data real de pagamento e uma linha de totais no
    fim. O chamador só precisa dar wb.save(caminho)."""
    from datetime import datetime as _dt
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from utils.helpers import PAYMENT_METHODS

    GREEN = "1F8A5B"
    RED   = "C0392B"
    HEADER_BG = "1F2937"

    wb = Workbook()
    ws = wb.active
    ws.title = (month_name or "Lançamentos")[:31]

    headers = ["Tipo", "Descrição", "Categoria", "Forma de Pagamento", "Valor", "Data", "Previsto"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    rows = get_transactions(month_id)
    total_entradas = 0.0
    total_saidas   = 0.0
    r_idx = 2
    for r in rows:
        is_income = r["type"] in ("entrada_fixa", "entrada_variavel")
        amount = float(r["amount"] or 0)
        raw_date = r.get("payment_date") or r.get("created_at")
        try:
            data_val = _dt.fromisoformat(str(raw_date)[:19]).date() if raw_date else None
        except ValueError:
            data_val = None

        values = [
            _EXPORT_TYPE_LABELS.get(r["type"], r["type"]),
            r["description"],
            r["category"] or "Outros",
            PAYMENT_METHODS.get(r.get("payment_method"), "") if r.get("payment_method") else "",
            amount,
            data_val,
            "Sim" if r.get("is_expectation") else "Não",
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c, value=val)
            cell.border = border
            if c == 5:
                cell.number_format = '"R$" #,##0.00'
                cell.font = Font(color=GREEN if is_income else RED, bold=True)
            elif c == 6 and data_val:
                cell.number_format = "DD/MM/YYYY"

        if not r.get("is_expectation"):
            if is_income:
                total_entradas += amount
            else:
                total_saidas += amount
        r_idx += 1

    # ── Linha de totais ─────────────────────────────────────────────────
    r_idx += 1
    for label, value, color in (
        ("Total Entradas", total_entradas, GREEN),
        ("Total Saídas",   total_saidas,   RED),
        ("Saldo",          total_entradas - total_saidas, None),
    ):
        ws.cell(row=r_idx, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=r_idx, column=5, value=value)
        cell.number_format = '"R$" #,##0.00'
        cell.font = Font(bold=True, color=color or ("1F2937"))
        r_idx += 1

    widths = [16, 38, 18, 18, 14, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return wb


# ---------------------------------------------------------------------------
# Configuração do Dashboard (widgets habilitados/ordem) — sincronizada com
# o site via user_settings.dashboard_widgets. Ao contrário do tema, não
# precisa de cache local: o Dashboard só existe depois do login.
# ---------------------------------------------------------------------------

def get_dashboard_widgets() -> Optional[list]:
    """Lista ordenada de {"id": ..., "enabled": ...}, ou None se o usuário
    nunca customizou (o chamador usa o catálogo padrão nesse caso)."""
    resp = get_client().table("user_settings").select("dashboard_widgets").execute()
    if not resp.data:
        return None
    return resp.data[0].get("dashboard_widgets")


def save_dashboard_widgets(config: list) -> None:
    client  = get_client()
    user_id = client.auth.get_user().user.id
    existing = client.table("user_settings").select("user_id").eq("user_id", user_id).execute()
    if existing.data:
        client.table("user_settings").update({"dashboard_widgets": config}).eq("user_id", user_id).execute()
    else:
        client.table("user_settings").insert({"user_id": user_id, "dashboard_widgets": config}).execute()


# ---------------------------------------------------------------------------
# Dia de corte da importação de extrato (config do usuário — mesma tabela
# user_settings, sincronizada com o site). Lançamentos a partir desse dia do
# mês contam pro mês seguinte. Padrão 1 (sem deslocamento) se o usuário
# nunca configurou, ou se a migration 018 ainda não rodou no Supabase.
# ---------------------------------------------------------------------------

def get_import_cutoff_day() -> int:
    from utils.helpers import DEFAULT_IMPORT_CUTOFF_DAY
    try:
        resp = get_client().table("user_settings").select("import_cutoff_day").execute()
    except Exception:
        return DEFAULT_IMPORT_CUTOFF_DAY
    if not resp.data or resp.data[0].get("import_cutoff_day") is None:
        return DEFAULT_IMPORT_CUTOFF_DAY
    return resp.data[0]["import_cutoff_day"]


def save_import_cutoff_day(day: int) -> None:
    client  = get_client()
    user_id = client.auth.get_user().user.id
    existing = client.table("user_settings").select("user_id").eq("user_id", user_id).execute()
    if existing.data:
        client.table("user_settings").update({"import_cutoff_day": day}).eq("user_id", user_id).execute()
    else:
        client.table("user_settings").insert({"user_id": user_id, "import_cutoff_day": day}).execute()
